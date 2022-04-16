"""
========================
Author:阿杰
Time:2022/3/16 12:57
E-mail:482564719@qq.com
========================
"""
import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        """

        :param filename: excel文件名
        :param sheetname: excel表单名
        """

        self.filename = filename
        self.sheetname = sheetname

    # 设置读取excel数据方法
    def read_data(self):
        """

        :return:
        """
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # rows按行获取表单中所有的格子,每一行的格子放到一个元组中
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行意外所有的行
        for item in res[1:]:
            # 获取该行的值
            data = [i.value for i in item]
            # 第一行的数据和当前这行数据打包为字典
            dic = dict(zip(title, data))
            # 把字典添加到cases这个列表中
            cases.append(dic)
        return cases

    # 设置写入excel数据方法
    def write_data(self, row, colum, value):
        """

        :param row: 写入的行
        :param colum: 写入的列
        :param value: 写入的值
        :return: 返回读出来的值
        """
        # 加载工作薄对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]

        # excel写入数据
        sh.cell(row=row, column=colum, value=value)
        workbook.save(self.filename)
